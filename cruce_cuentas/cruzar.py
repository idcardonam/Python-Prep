#!/usr/bin/env python3
"""
Cruce reutilizable: Google Workspace + información académica (+ personal opcional).

Uso:
  python3 cruzar.py --ejemplo
  python3 cruzar.py --google entrada/google_admin.csv --academico entrada/academico.csv
  python3 cruzar.py --google g.csv --academico a.csv --personal p.csv --salida salida/

No subas CSV reales a Git. Solo corre en tu PC.
"""
from __future__ import annotations

import argparse
import csv
import html
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import yaml  # type: ignore
except ImportError:
    yaml = None


# --- Detección de columnas (exportes Admin Console EN/ES) ---

GOOGLE_ALIASES = {
    "email": [
        "email address [required]",
        "email address",
        "email",
        "correo electrónico",
        "correo",
        "primary email",
    ],
    "nombre": ["first name [required]", "first name", "nombre", "given name"],
    "apellido": ["last name [required]", "last name", "apellidos", "apellido", "family name"],
    "estado_google": ["status [read only]", "status", "estado", "account status"],
    "ou": ["org unit path [required]", "org unit path", "org unit", "unidad organizacional", "ou"],
    "ultimo_ingreso": [
        "last sign in [read only]",
        "last sign in",
        "last login",
        "último inicio de sesión",
        "ultimo inicio de sesion",
    ],
    "2fa_inscrito": [
        "2sv enrolled [read only]",
        "2sv enrolled",
        "2-step verification enrollment status",
        "2sv enrollment",
        "enrollment status",
        "verificación en 2 pasos",
        "2-step verification enrolled",
    ],
    "2fa_forzado": [
        "2-step verification enforcement",
        "2sv enforcement",
        "enforcement",
        "aplicación de verificación en 2 pasos",
    ],
}

ACADEMICO_ALIASES = {
    "email": [
        "correo_unab",
        "correo institucional",
        "correo_institucional",
        "email institucional",
        "correo unab",
        "mail_institucional",
        "correo",
        "email",
        "e-mail",
        "mail",
        "usuario",
        "userprincipalname",
        "cuenta",
        "e_mail",
    ],
    "estado": [
        "est_acad",
        "estado academico",
        "estado académico",
        "estado",
        "situacion",
        "situación",
    ],
    "facultad": ["escuela", "facultad", "unidad academica", "unidad académica"],
    "programa": ["programa", "carrera", "programa academico", "programa académico"],
    "seccion": ["seccion", "sección", "paralelo"],
    "jornada": ["jornada", "modalidad jornada"],
    "codigo": ["codigo", "código", "codigo_estudiante", "codigo estudiante", "id estudiante", "id"],
    "nivel": ["nivel", "tipo formacion", "tipo formación", "nivel formacion"],
    "nombres": ["nombres", "nombre", "primer nombre"],
    "apellidos": ["apellidos", "apellido"],
    "cod_prog": ["cod_prog"],
    "cod_majr": ["cod_majr"],
    "cod_esc": ["cod_esc"],
    "periodo": ["periodo"],
}

PERSONAL_ALIASES = {
    "email": ["correo", "email", "correo institucional"],
    "tipo": ["tipo", "vinculacion", "vinculación", "rol", "categoria"],
    "area": ["area", "área", "dependencia", "facultad", "unidad"],
    "seccion": ["seccion", "sección", "oficina", "grupo"],
    "cargo": ["cargo", "puesto"],
    "nombres": ["nombres", "nombre"],
    "apellidos": ["apellidos", "apellido"],
}

CURRICULO_ALIASES = {
    "periodo": ["term", "periodo", "periodo_plan", "effective term"],
    "periodo_efectivo": ["term_eff", "effective term"],
    "escuela": ["desc_escuela", "escuela", "facultad", "college"],
    "cod_esc": ["cod_escuela", "cod_esc", "codigo_escuela"],
    "programa": ["desc_prog", "programa", "carrera", "program"],
    "cod_prog": ["cod_prog", "codigo_programa"],
    "major": ["cod_majr", "major", "cod_major"],
    "nivel": ["nivel_desc", "nivel"],
    "plan": ["rule_curr", "plan", "plan_estudios", "plan estudios", "curriculo", "cod_plan"],
    "tipo": ["tipo", "formal", "tipo_formacion", "tipo formación", "modalidad"],
    "campus": ["desc_camp", "campus"],
    "distrito": ["desc_distrito", "distrito"],
    "web": ["web_ind"],
}

EXCLUIR_DEFAULT = {
    "GRADUADO",
    "GRADUADA",
    "EGRESADO",
    "EGRESADA",
    "TITULADO",
    "TITULADA",
    "GRADUATE",
    "ALUMNI",
}


def norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    return re.sub(r"\s+", " ", s)


def norm_email(s: str) -> str:
    return (s or "").strip().lower()


def norm_estado(s: str) -> str:
    s = (s or "").strip().upper()
    s = s.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    return re.sub(r"\s+", " ", s)


def pick_col(headers: list[str], aliases: list[str]) -> str | None:
    """Coincidencia exacta del encabezado normalizado. Evita CANT_CURSOS / INSCRIP_STATUS."""
    mapped = {norm_header(h): h for h in headers}
    for a in aliases:
        hit = mapped.get(norm_header(a))
        if hit:
            return hit
    return None


def map_columns(headers: list[str], alias_map: dict[str, list[str]]) -> dict[str, str | None]:
    return {k: pick_col(headers, v) for k, v in alias_map.items()}


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    first = next((ln for ln in text.splitlines() if ln.strip()), "")
    n_comma, n_semi = first.count(","), first.count(";")
    if n_semi > n_comma:
        delimiter = ";"
    elif n_comma > n_semi:
        delimiter = ","
    else:
        try:
            delimiter = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|").delimiter
        except csv.Error:
            delimiter = ","
    reader = csv.DictReader(text.splitlines(), delimiter=delimiter)
    headers = reader.fieldnames or []
    rows = [{k: (v if v is not None else "").strip() for k, v in row.items() if k is not None} for row in reader]
    headers = [h for h in headers if h is not None]
    if len(headers) <= 1 and n_semi > 1:
        reader = csv.DictReader(text.splitlines(), delimiter=";")
        headers = [h for h in (reader.fieldnames or []) if h is not None]
        rows = [{k: (v if v is not None else "").strip() for k, v in row.items() if k is not None} for row in reader]
    return headers, rows


def _celda(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def read_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("Para leer .xlsx instala openpyxl:  py -3 -m pip install openpyxl") from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == "sql":
            continue
        ws = wb[name]
        break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    raw_h = next(it, None) or []
    headers = [str(h).strip() if h is not None and str(h).strip() else f"col_{i}" for i, h in enumerate(raw_h)]
    rows: list[dict[str, str]] = []
    for row in it:
        d = {headers[i]: _celda(row[i] if i < len(row) else None) for i in range(len(headers))}
        if any(d.values()):
            rows.append(d)
    wb.close()
    return headers, rows


def read_tabla(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    return read_csv(path)


def periodo_desde_nombre(nombre: str) -> str:
    m = re.search(r"(20\d{4})", nombre)
    if m:
        return m.group(1)
    m = re.search(r"IC\s*20\d{2}", nombre, re.I)
    return re.sub(r"\s+", " ", m.group(0)).upper() if m else Path(nombre).stem


def parece_export_google(headers: list[str]) -> bool:
    blob = " ".join(norm_header(h) for h in headers)
    return "2sv" in blob or "2-step verification" in blob or "email address [required]" in blob


def nombre_parece_curriculo(path: Path) -> bool:
    nom = norm_header(path.stem.replace("-", " ").replace("_", " "))
    return "curricul" in nom or "vista de curriculo" in nom


def es_archivo_curriculo(path: Path, headers: list[str] | None = None) -> bool:
    if nombre_parece_curriculo(path):
        return True
    h = headers if headers is not None else peek_headers(path)
    nh = {norm_header(x) for x in h}
    if "correo_unab" in nh or "correo institucional" in nh:
        return False
    if parece_export_google(h):
        return False
    return bool(nh & {"term_eff", "term", "periodo"}) and bool(
        nh & {"desc_escuela", "cod_escuela", "cod_prog", "desc_prog", "programa", "escuela"}
    )


def peek_headers(path: Path) -> list[str]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        headers, _ = read_xlsx(path)
        return headers
    raw = path.read_bytes()[:16384]
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    first = next((ln for ln in text.splitlines() if ln.strip()), "")
    delim = ";" if first.count(";") > first.count(",") else ","
    return next(csv.reader([first], delimiter=delim), [])


def listar_csv(carpeta: Path) -> list[Path]:
    return sorted(p for p in carpeta.glob("*.csv") if p.is_file())


def listar_tablas(carpeta: Path) -> list[Path]:
    return sorted(
        p
        for p in carpeta.iterdir()
        if p.is_file() and p.suffix.lower() in {".csv", ".xlsx", ".xlsm"}
    )


def listar_csv_academico(carpeta: Path) -> list[Path]:
    out: list[Path] = []
    for p in listar_csv(carpeta):
        if p.name.lower().startswith("user_download"):
            continue
        headers = peek_headers(p)
        if parece_export_google(headers) or es_archivo_curriculo(p, headers):
            continue
        nh = {norm_header(x) for x in headers}
        if {"tipo", "cargo"} <= nh and "escuela" not in nh and "est_acad" not in nh:
            continue
        out.append(p)
    return out


def listar_csv_curriculo(carpeta: Path) -> list[Path]:
    return [p for p in listar_tablas(carpeta) if es_archivo_curriculo(p)]


def exigir_carpeta(carpeta: Path) -> Path:
    """La carpeta de --carpeta es la de TUS CSV, no la del repo Python-Prep."""
    p = Path(carpeta)
    if p.is_dir():
        return p
    raise SystemExit(
        "No existe esa carpeta:\n"
        f"  {p}\n\n"
        "Eso no es un error de la rama. --carpeta debe ser la ruta REAL donde están:\n"
        "  - User_Download*.csv  (Google Admin)\n"
        "  - Prematriculados / inscritos *.csv\n"
        "  - VISTA DE CURRICULO.xlsx (si lo tienes)\n\n"
        "En el Explorador de Windows abre esa carpeta, clic en la barra de dirección,\n"
        "copia la ruta y pégala entre comillas en --carpeta.\n"
        "Si ya estás en cruce_cuentas, no vuelvas a hacer: cd cruce_cuentas"
    )


def descubrir_fuentes(carpeta: Path) -> dict[str, list[Path]]:
    """Una carpeta con Google + inscritos + currículo. Elige el Google más reciente."""
    carpeta = exigir_carpeta(carpeta)
    google: list[Path] = []
    curric: list[Path] = []
    for p in listar_tablas(carpeta):
        h = peek_headers(p)
        if p.name.lower().startswith("user_download") or parece_export_google(h):
            google.append(p)
        elif es_archivo_curriculo(p, h):
            curric.append(p)
    google.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    curric.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return {"google": google, "curriculo": curric, "academico": listar_csv_academico(carpeta)}


def inspeccionar_academico(carpeta: Path) -> None:
    """Solo metadatos: no imprime correos ni celdas."""
    carpeta = exigir_carpeta(carpeta)
    todos = listar_tablas(carpeta)
    google = [p for p in todos if p.name.lower().startswith("user_download") or parece_export_google(peek_headers(p))]
    curric = listar_csv_curriculo(carpeta)
    files = listar_csv_academico(carpeta)
    print(f"Carpeta: {carpeta}")
    if google:
        print(f"Google Admin omitido del académico: {len(google)}")
        for p in google:
            print(f"  - {p.name}")
    if curric:
        print(f"Vista de currículo (catálogo de planes, no estudiantes): {len(curric)}")
        for p in curric:
            headers, rows = read_tabla(p)
            mapped = map_columns(headers, CURRICULO_ALIASES)
            print(f"  {p.name}  filas={len(rows)}  cols={len(headers)}")
            print(f"    periodo: {mapped.get('periodo') or 'NO'}  escuela: {mapped.get('escuela') or 'NO'}  programa: {mapped.get('programa') or 'NO'}")
            print(f"    cod_prog: {mapped.get('cod_prog') or 'NO'}  major: {mapped.get('major') or 'NO'}  plan: {mapped.get('plan') or 'NO'}")
            print(f"    columnas: {headers}")
    if not files:
        if curric:
            print("No hay CSV de inscritos; sí hay currículo (catálogo).")
            return
        raise SystemExit(f"No hay CSV académicos de inscritos en {carpeta}")
    print(f"Archivos CSV de inscritos: {len(files)}")
    print("-" * 72)
    all_headers: dict[str, set[str]] = {}
    total = 0
    for p in files:
        headers, rows = read_csv(p)
        total += len(rows)
        mapped = map_columns(headers, ACADEMICO_ALIASES)
        print(f"{p.name}")
        print(f"  filas={len(rows):>8}  cols={len(headers):>3}  periodo={periodo_desde_nombre(p.name)}  bytes={p.stat().st_size}")
        print(f"  correo detectado: {mapped.get('email') or 'NO'}")
        print(f"  estado detectado: {mapped.get('estado') or 'NO'}")
        print(f"  facultad: {mapped.get('facultad') or 'NO'}  programa: {mapped.get('programa') or 'NO'}  seccion: {mapped.get('seccion') or 'NO'}")
        print(f"  columnas: {headers}")
        print()
        all_headers.setdefault("union", set()).update(headers)
    print("-" * 72)
    print(f"Filas totales inscritos (suma bruta, puede haber duplicados entre archivos): {total}")
    print(f"Columnas distintas en inscritos: {len(all_headers['union'])}")
    print("Regla de currículo: si hay varias filas del mismo programa/major, el ÚLTIMO periodo es el plan vigente.")


def cargar_academico_carpeta(carpeta: Path) -> tuple[list[str], list[dict[str, str]]]:
    files = listar_csv_academico(carpeta)
    if not files:
        raise SystemExit(f"No hay CSV en {carpeta}")
    union: list[str] = []
    seen_h: set[str] = set()
    rows_out: list[dict[str, str]] = []
    for p in files:
        headers, rows = read_csv(p)
        for h in headers:
            if h not in seen_h:
                seen_h.add(h)
                union.append(h)
        periodo = periodo_desde_nombre(p.name)
        for row in rows:
            row["_fuente_archivo"] = p.name
            row["_periodo"] = periodo
            rows_out.append(row)
    extra = ["_fuente_archivo", "_periodo"]
    headers_final = union + [x for x in extra if x not in seen_h]
    return headers_final, rows_out


def merge_texto(prev: str, nuevo: str) -> str:
    prev = (prev or "").strip()
    nuevo = (nuevo or "").strip()
    if not nuevo:
        return prev
    if not prev:
        return nuevo
    parts = [x.strip() for x in prev.split(" | ") if x.strip()]
    if nuevo not in parts:
        parts.append(nuevo)
    return " | ".join(parts)


def split_multi(s: str) -> list[str]:
    return [x.strip() for x in (s or "").replace(";", "|").split("|") if x.strip()]


def periodo_orden(valor: str) -> tuple[int, int]:
    """Mayor tupla = periodo más reciente (plan vigente)."""
    s = (valor or "").strip().upper()
    m = re.search(r"(20\d{4})", s)
    if m:
        return (2, int(m.group(1)))
    m = re.search(r"IC\s*(20\d{2})", s)
    if m:
        return (1, int(m.group(1)))
    m = re.search(r"(\d{5,6})", s)
    if m:
        return (2, int(m.group(1)))
    return (0, 0)


def clave_txt(s: str) -> str:
    return norm_header(s).upper()


def indice_curriculo(cod_prog: str, major: str, programa: str, escuela: str) -> list[tuple]:
    keys: list[tuple] = []
    cp, cm = clave_txt(cod_prog), clave_txt(major)
    pr, es = clave_txt(programa), clave_txt(escuela)
    if cp and cm:
        keys.append(("pm", cp, cm))
    if cp:
        keys.append(("p", cp))
    if pr:
        keys.append(("n", pr, es))
    return keys


def cargar_catalogo_curriculo(paths: list[Path]) -> tuple[dict[tuple, dict[str, str]], dict[str, str | None], int]:
    """Una fila vigente por programa/major: gana el último PERIODO."""
    vigentes: dict[tuple, dict[str, str]] = {}
    c_map: dict[str, str | None] = {}
    n_hist = 0
    for path in paths:
        headers, rows = read_tabla(path)
        c_map = map_columns(headers, CURRICULO_ALIASES)
        for row in rows:
            n_hist += 1
            pieza = {campo: (row.get(c_map[campo], "") if c_map.get(campo) else "") for campo in c_map}
            pieza["_fuente"] = path.name
            pieza["_todas"] = dict(row)
            per = pieza.get("periodo") or ""
            orden = periodo_orden(per)
            for key in indice_curriculo(pieza.get("cod_prog") or "", pieza.get("major") or "", pieza.get("programa") or "", pieza.get("escuela") or ""):
                prev = vigentes.get(key)
                if prev is None or orden >= periodo_orden(prev.get("periodo") or ""):
                    vigentes[key] = pieza
    return vigentes, c_map, n_hist


def buscar_plan(catalogo: dict[tuple, dict[str, str]], cod_prog: str, major: str, programa: str, escuela: str) -> dict[str, str]:
    for key in indice_curriculo(cod_prog, major, programa, escuela):
        hit = catalogo.get(key)
        if hit:
            return hit
    cps, cms, prs, ess = split_multi(cod_prog), split_multi(major), split_multi(programa), split_multi(escuela)
    n = max(len(cps), len(cms), len(prs), len(ess), 1)
    for i in range(n):
        hit = buscar_plan_simple(
            catalogo,
            cps[i] if i < len(cps) else (cps[-1] if cps else ""),
            cms[i] if i < len(cms) else (cms[-1] if cms else ""),
            prs[i] if i < len(prs) else (prs[-1] if prs else ""),
            ess[i] if i < len(ess) else (ess[-1] if ess else ""),
        )
        if hit:
            return hit
    return {}


def buscar_plan_simple(catalogo: dict[tuple, dict[str, str]], cod_prog: str, major: str, programa: str, escuela: str) -> dict[str, str]:
    for key in indice_curriculo(cod_prog, major, programa, escuela):
        hit = catalogo.get(key)
        if hit:
            return hit
    return {}


def adjuntar_curriculo(ficha: dict[str, Any], plan: dict[str, str]) -> None:
    if not plan:
        return
    dest = ficha.setdefault("curriculo", {})
    for campo in ("periodo", "periodo_efectivo", "escuela", "cod_esc", "programa", "cod_prog", "major", "nivel", "plan", "tipo", "campus", "distrito", "web"):
        dest[campo] = merge_texto(dest.get(campo, ""), plan.get(campo, ""))
    dest["match"] = "SI"


def parse_date(value: str) -> datetime | None:
    v = (value or "").strip()
    if not v or v in {"-", "Never", "Nunca", "—", "Never logged in"}:
        return None
    v = v.replace("T", " ")
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(v[:19], fmt)
        except ValueError:
            continue
    return None


def is_2fa_on(inscrito: str, forzado: str) -> bool:
    """True solo si el usuario tiene 2FA inscrito. Enforcement Off/On es otra cosa."""
    i = (inscrito or "").strip().lower()
    if not i:
        return False
    if i in {"not enrolled", "no", "off", "false", "0", "never", "no inscrito", "not_enrolled"}:
        return False
    if "not enrolled" in i or i.startswith("not "):
        return False
    if i in {"enrolled", "yes", "si", "sí", "true", "on", "inscrito"}:
        return True
    if "enrolled" in i or "inscrit" in i:
        return True
    return False


# Textos del HTML. Cámbialos aquí o en config.yaml → informe:
TITULOS_INFORME = {
    "marca": "UNAB · Dirección de TIC",
    "h1": "Estudiantes vigentes sin 2FA",
    "intro": "Google Admin (2FA inscrito) × inscritos Banner × plan vigente de currículo.",
    "resumen": "Resumen",
    "texto_resumen": "Solo estudiantes con matrícula vigente y cuenta @unab.edu.co. Egresados y personal no entran aquí.",
    "kpi_universo": "Vigentes con Google",
    "kpi_con": "Con 2FA",
    "kpi_sin": "Pendientes",
    "kpi_cob": "Cobertura",
    "facultades": "Facultades",
    "programas": "Programas",
    "correos": "Correos",
    "listado_h1": "Correos sin 2FA",
    "volver_informe": "Volver al informe",
}


def load_config(path: Path | None) -> dict[str, Any]:
    cfg: dict[str, Any] = {
        "dominio": "unab.edu.co",
        "estados_excluir": sorted(EXCLUIR_DEFAULT),
        "ou_personal": ["docente", "profesor", "administrativo", "gestion", "tic", "staff", "planta"],
        "dias_inactiva": 90,
        "informe": dict(TITULOS_INFORME),
    }
    if path and path.exists() and yaml:
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        inf = data.get("informe")
        cfg.update({k: v for k, v in data.items() if v is not None and k != "informe"})
        if isinstance(inf, dict):
            merged = dict(TITULOS_INFORME)
            merged.update({str(k): str(v) for k, v in inf.items() if v is not None})
            cfg["informe"] = merged
    return cfg


def titulos_informe(cfg: dict[str, Any] | None) -> dict[str, str]:
    t = dict(TITULOS_INFORME)
    extra = (cfg or {}).get("informe") or {}
    if isinstance(extra, dict):
        t.update({str(k): str(v) for k, v in extra.items() if v is not None})
    return t


def proyecto_a_ficha(email: str) -> dict[str, Any]:
    return {
        "correo": email,
        "perfil": "SIN_CLASIFICAR",
        "prioridad_2fa": "N/A",
        "tiene_2fa": None,
        "en_google": False,
        "en_academico": False,
        "en_personal": False,
        "google": {},
        "academico": {},
        "personal": {},
        "curriculo": {},
        "alertas": [],
    }


def clasificar(ficha: dict[str, Any], ou_personal: list[str], excluir: set[str]) -> None:
    ou = (ficha["google"].get("ou") or "").lower()
    estado = norm_estado(ficha["academico"].get("estado") or "")
    if estado in excluir:
        ficha["perfil"] = "EGRESADO_EN_EXTRACTO"
        ficha["alertas"].append("Aparece en académico con estado de egreso (debería filtrarse).")
        return
    if ficha["en_academico"] and ficha["en_google"]:
        ficha["perfil"] = "ESTUDIANTE_VIGENTE"
        return
    if ficha["en_academico"] and not ficha["en_google"]:
        ficha["perfil"] = "ACADEMICO_SIN_CUENTA_GOOGLE"
        ficha["alertas"].append("Tiene registro académico vigente y no aparece en Admin Google.")
        return
    if ficha["en_personal"]:
        tipo = (ficha["personal"].get("tipo") or "").upper()
        ficha["perfil"] = "PERSONAL_" + (tipo if tipo else "INSTITUCIONAL")
        return
    if any(p in ou for p in ou_personal):
        ficha["perfil"] = "POSIBLE_PERSONAL_POR_OU"
        return
    if "egresad" in ou or "gradua" in ou or "alumni" in ou:
        ficha["perfil"] = "POSIBLE_EGRESADO_CON_CUENTA"
        ficha["alertas"].append("Está en Google y no en académico vigente: posible egresado con cuenta activa.")
        return
    if ficha["en_google"]:
        ficha["perfil"] = "GOOGLE_SIN_MATCH_ACADEMICO"
        ficha["alertas"].append(
            "Cuenta institucional sin ficha académica vigente: personal, egresado u órfana. Completar con CSV de GH."
        )
        return
    ficha["perfil"] = "SIN_CLASIFICAR"


def prioridad_2fa(ficha: dict[str, Any], dias_inactiva: int, now: datetime) -> str:
    if ficha["tiene_2fa"] is True:
        return "OK"
    if ficha["tiene_2fa"] is None:
        return "SIN_DATO_GOOGLE"
    ultimo = parse_date(ficha["google"].get("ultimo_ingreso") or "")
    inactiva = False
    if ultimo is None:
        inactiva = True
    else:
        inactiva = (now - ultimo).days >= dias_inactiva
    perfil = ficha["perfil"]
    if perfil.startswith("PERSONAL") or perfil == "POSIBLE_PERSONAL_POR_OU":
        return "ALTA_PERSONAL_SIN_2FA"
    if perfil == "ESTUDIANTE_VIGENTE" and not inactiva:
        return "ALTA_ESTUDIANTE_ACTIVO_SIN_2FA"
    if perfil == "ESTUDIANTE_VIGENTE" and inactiva:
        return "MEDIA_ESTUDIANTE_INACTIVO_SIN_2FA"
    if perfil == "POSIBLE_EGRESADO_CON_CUENTA":
        return "REVISAR_EGRESO_Y_2FA"
    if inactiva:
        return "MEDIA_CUENTA_INACTIVA_SIN_2FA"
    return "ALTA_SIN_2FA"


def salida_escribible(salida: Path) -> Path:
    """Si Excel dejó un CSV abierto, Windows bloquea .\\salida; usa una carpeta nueva."""
    salida.mkdir(parents=True, exist_ok=True)
    probe = salida / "_probe_escritura.tmp"
    try:
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
        for n in (
            "00_universo.csv",
            "07_cobertura_2fa_programa_seccion.csv",
            "resumen.html",
            "catalogo_facultades.html",
        ):
            p = salida / n
            if p.exists():
                with p.open("a", encoding="utf-8"):
                    pass
        return salida
    except OSError:
        alt = Path(str(salida) + "_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
        alt.mkdir(parents=True, exist_ok=True)
        print("AVISO: no pude escribir en", salida.resolve())
        print("Cierra Excel/CSV/HTML de esa carpeta. Esta corrida queda en:")
        print(" ", alt.resolve())
        return alt


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fieldnames})


def flatten(ficha: dict[str, Any]) -> dict[str, Any]:
    g, a, p = ficha["google"], ficha["academico"], ficha["personal"]
    c = ficha.get("curriculo") or {}
    out = {
        "correo": ficha["correo"],
        "perfil": ficha["perfil"],
        "prioridad_2fa": ficha["prioridad_2fa"],
        "tiene_2fa": "" if ficha["tiene_2fa"] is None else ("SI" if ficha["tiene_2fa"] else "NO"),
        "en_google": "SI" if ficha["en_google"] else "NO",
        "en_academico": "SI" if ficha["en_academico"] else "NO",
        "en_personal": "SI" if ficha["en_personal"] else "NO",
        "g_nombre": g.get("nombre", ""),
        "g_apellido": g.get("apellido", ""),
        "g_estado": g.get("estado_google", ""),
        "g_ou": g.get("ou", ""),
        "g_ultimo_ingreso": g.get("ultimo_ingreso", ""),
        "g_2fa_inscrito": g.get("2fa_inscrito", ""),
        "g_2fa_forzado": g.get("2fa_forzado", ""),
        "a_nombres": a.get("nombres", ""),
        "a_apellidos": a.get("apellidos", ""),
        "a_estado": a.get("estado", ""),
        "a_facultad": a.get("facultad", ""),
        "a_programa": a.get("programa", ""),
        "a_seccion": a.get("seccion", ""),
        "a_jornada": a.get("jornada", ""),
        "a_codigo": a.get("codigo", ""),
        "a_nivel": a.get("nivel", ""),
        "a_cod_prog": a.get("cod_prog", ""),
        "a_cod_majr": a.get("cod_majr", ""),
        "c_match": c.get("match", "NO") if c else "NO",
        "c_periodo_vigente": c.get("periodo", ""),
        "c_escuela": c.get("escuela", ""),
        "c_programa": c.get("programa", ""),
        "c_cod_prog": c.get("cod_prog", ""),
        "c_major": c.get("major", ""),
        "c_plan": c.get("plan", ""),
        "c_nivel": c.get("nivel", ""),
        "c_tipo": c.get("tipo", ""),
        "c_campus": c.get("campus", ""),
        "c_distrito": c.get("distrito", ""),
        "c_web": c.get("web", ""),
        "c_periodo_efectivo": c.get("periodo_efectivo", ""),
        "p_tipo": p.get("tipo", ""),
        "p_area": p.get("area", ""),
        "p_seccion": p.get("seccion", ""),
        "p_cargo": p.get("cargo", ""),
        "alertas": " | ".join(ficha["alertas"]),
    }
    extra = a.get("_todas") or {}
    skip = {norm_header(x) for x in (
        "correo", "email", "correo institucional", "e-mail", "mail",
        "nombres", "nombre", "apellidos", "apellido", "estado",
        "facultad", "programa", "seccion", "sección", "jornada",
        "codigo", "código", "codigo_estudiante", "nivel",
        "cod_prog", "cod_majr", "cod_esc", "periodo",
    )}
    for k, v in extra.items():
        if not k or norm_header(k) in skip:
            continue
        out[f"a_{k}"] = v
    return out


def familia_nivel(nivel: str) -> str:
    n = clave_txt(nivel)
    if any(x in n for x in ("NO FORMAL", "COMPLEMENTARIA")):
        return "Educación continua / no formal"
    if any(x in n for x in ("BACHILLER", "PRIMARIA", "PRE-ESCOLAR", "PRE ESCOLAR")):
        return "Instituto Caldas / básica"
    if any(x in n for x in ("TECNICO LABORAL", "PREGRADO TECNICO", "PREGRADO TECNOLOGIA", "TECNICO AVANZADO", "ESPECIALIZACION TECNOLOGICA")):
        return "Técnico y tecnológico"
    if any(x in n for x in ("MAESTR", "DOCTOR", "ESPECIALIZ", "POSGRADO", "COTERMINAL")):
        return "Posgrado"
    if "PREGRADO" in n or "PROFESIONAL" in n:
        return "Pregrado profesional"
    return "Otro"


def slug_id(s: str) -> str:
    s = clave_txt(s)
    s = re.sub(r"[^A-Z0-9]+", "-", s).strip("-")
    return s[:80] or "x"


def html_catalogo_facultades(path: Path, planes: list[dict[str, Any]], generado: str) -> None:
    """Página aparte: oferta vigente agrupada por tipo y facultad (lo que pidió el ing)."""
    rows = []
    for r in planes:
        item = dict(r)
        item["familia"] = familia_nivel(str(r.get("nivel") or ""))
        rows.append(item)
    orden_fam = [
        "Pregrado profesional",
        "Posgrado",
        "Técnico y tecnológico",
        "Instituto Caldas / básica",
        "Educación continua / no formal",
        "Otro",
    ]
    fam_count = Counter(r["familia"] for r in rows)
    nav = "".join(
        f'<a href="#fam-{slug_id(f)}">{html.escape(f)} ({fam_count[f]})</a>'
        for f in orden_fam
        if fam_count[f]
    )
    bloques: list[str] = []
    for fam in orden_fam:
        grupo = [r for r in rows if r["familia"] == fam]
        if not grupo:
            continue
        by_esc: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for r in grupo:
            by_esc[str(r.get("escuela") or "(sin escuela)")].append(r)
        tarjetas = "".join(
            f'<a class="cardlet" href="#esc-{slug_id(fam + esc)}"><strong>{html.escape(esc)}</strong>'
            f"<span>{len(filas)} programas</span></a>"
            for esc, filas in sorted(by_esc.items(), key=lambda x: x[0])
        )
        inner: list[str] = []
        for esc, filas in sorted(by_esc.items(), key=lambda x: x[0]):
            filas = sorted(filas, key=lambda r: (str(r.get("nivel") or ""), str(r.get("programa") or "")))
            tr = "".join(
                "<tr>"
                f"<td>{html.escape(str(r.get('programa') or ''))}</td>"
                f"<td>{html.escape(str(r.get('cod_prog') or ''))}</td>"
                f"<td>{html.escape(str(r.get('major') or ''))}</td>"
                f"<td>{html.escape(str(r.get('nivel') or ''))}</td>"
                f"<td>{html.escape(str(r.get('periodo_vigente') or ''))}</td>"
                f"<td>{html.escape(str(r.get('campus') or ''))}</td>"
                "</tr>"
                for r in filas
            )
            inner.append(
                f'<section class="fac" id="esc-{slug_id(fam + esc)}"><h3>{html.escape(esc)}</h3>'
                f"<p class=\"note\">{len(filas)} planes vigentes (último TERM)</p>"
                "<table><thead><tr><th>Programa</th><th>Cód.</th><th>Major</th><th>Nivel</th>"
                f"<th>Último TERM</th><th>Campus</th></tr></thead><tbody>{tr}</tbody></table></section>"
            )
        bloques.append(
            f'<section class="familia" id="fam-{slug_id(fam)}"><h2>{html.escape(fam)}</h2>'
            f'<div class="grid">{tarjetas}</div>{"".join(inner)}</section>'
        )
    doc = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"/>
<title>Oferta por facultades — UNAB</title>
<style>
body{{font-family:Segoe UI,Arial,sans-serif;margin:0;background:#f4f7fb;color:#1f2937}}
header{{background:#003B70;color:#fff;padding:1.4rem 2rem}}
header a{{color:#fff}}
nav{{display:flex;flex-wrap:wrap;gap:.6rem;padding:1rem 2rem;background:#e8eef5;position:sticky;top:0}}
nav a{{background:#fff;padding:.35rem .7rem;border-radius:999px;text-decoration:none;color:#003B70;font-size:.9rem}}
main{{padding:1rem 2rem 3rem}}
.grid{{display:flex;flex-wrap:wrap;gap:.6rem;margin:1rem 0}}
.cardlet{{background:#fff;border-radius:10px;padding:.7rem 1rem;min-width:180px;text-decoration:none;color:#1f2937;box-shadow:0 2px 8px rgba(0,0,0,.06)}}
.cardlet span{{display:block;color:#6b7280;font-size:.85rem}}
.fac{{background:#fff;border-radius:12px;padding:1rem;margin:1rem 0;box-shadow:0 2px 8px rgba(0,0,0,.05)}}
table{{border-collapse:collapse;width:100%;font-size:.92rem}}
th,td{{border-bottom:1px solid #d5deea;padding:.4rem;text-align:left}}
h1,h2,h3{{color:#003B70}} h1{{color:#fff;margin:0}}
.note{{color:#4b5563}}
</style></head><body>
<header>
<h1>Oferta vigente por facultad</h1>
<p>Vista de currículo · un programa = último TERM · {len(rows)} planes · {generado}</p>
<p><a href="resumen.html">Volver al cruce 2FA</a></p>
</header>
<nav>{nav or "Sin catálogo"}</nav>
<main>
<p class="note">Esto no es el listado de estudiantes. Es el catálogo de carreras/planes para organizar facultades. Educación continua queda en su bloque, no mezclada con pregrado.</p>
{"".join(bloques) or "<p>No se cargó VISTA DE CURRICULO. Pon el xlsx en la carpeta y vuelve a correr cruzar.py.</p>"}
</main></body></html>"""
    path.write_text(doc, encoding="utf-8")


def escuela_de(r: dict[str, Any]) -> str:
    return (r.get("c_escuela") or r.get("a_facultad") or "").strip() or "(sin facultad en currículo)"


def programa_de(r: dict[str, Any]) -> str:
    return (r.get("c_programa") or r.get("a_programa") or "").strip() or "(sin programa)"


def etiqueta_ingreso(valor: str) -> str:
    t = (valor or "").strip()
    if not t or t.lower().startswith("never"):
        return "Nunca ingresó"
    return t


def nombre_completo(r: dict[str, Any]) -> str:
    n = (r.get("a_nombres") or r.get("g_nombre") or "").strip()
    a = (r.get("a_apellidos") or r.get("g_apellido") or "").strip()
    return f"{n} {a}".strip()


def stats_campana(planos: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Counter], dict[tuple[str, str], Counter]]:
    vigentes = [r for r in planos if r.get("perfil") == "ESTUDIANTE_VIGENTE"]
    fac: dict[str, Counter] = defaultdict(Counter)
    prog: dict[tuple[str, str], Counter] = defaultdict(Counter)
    for r in vigentes:
        e, p = escuela_de(r), programa_de(r)
        fac[e]["cuentas"] += 1
        prog[(e, p)]["cuentas"] += 1
        if r.get("tiene_2fa") == "NO":
            fac[e]["sin"] += 1
            prog[(e, p)]["sin"] += 1
        if r.get("tiene_2fa") == "SI":
            fac[e]["con"] += 1
            prog[(e, p)]["con"] += 1
    return vigentes, fac, prog


def filas_campana_csv(planos: list[dict[str, Any]]) -> list[dict[str, str]]:
    filas = []
    for r in planos:
        if r.get("perfil") != "ESTUDIANTE_VIGENTE" or r.get("tiene_2fa") != "NO":
            continue
        filas.append(
            {
                "facultad": escuela_de(r),
                "programa": programa_de(r),
                "correo": str(r.get("correo") or ""),
                "nombres": str(r.get("a_nombres") or r.get("g_nombre") or ""),
                "apellidos": str(r.get("a_apellidos") or r.get("g_apellido") or ""),
                "estado_academico": str(r.get("a_estado") or ""),
                "ultimo_ingreso_google": etiqueta_ingreso(str(r.get("g_ultimo_ingreso") or "")),
                "periodo_plan_vigente": str(r.get("c_periodo_vigente") or ""),
                "cod_prog": str(r.get("c_cod_prog") or r.get("a_cod_prog") or ""),
            }
        )
    filas.sort(key=lambda x: (x["facultad"], x["programa"], x["correo"]))
    return filas


def css_presentacion() -> str:
    return """
:root { --azul:#003B70; --rojo:#b42318; --fondo:#f3f6fb; --texto:#1f2937; }
* { box-sizing: border-box; }
body { font-family: Segoe UI, Calibri, Arial, sans-serif; margin:0; background:var(--fondo); color:var(--texto); }
header { background:var(--azul); color:#fff; padding:1.6rem 2rem 1.3rem; }
header p { margin:.35rem 0 0; opacity:.92; max-width:900px; line-height:1.4; }
header a { color:#fff; }
.marca { font-size:.78rem; letter-spacing:.08em; text-transform:uppercase; opacity:.85; margin:0 0 .4rem; }
main { padding:1.2rem 2rem 3rem; max-width:1180px; }
.kpi { display:grid; grid-template-columns:repeat(auto-fit,minmax(170px,1fr)); gap:.8rem; margin:1rem 0 1.2rem; }
.kpi div { background:#fff; padding:1rem 1.1rem; border-radius:12px; box-shadow:0 2px 10px rgba(0,0,0,.06); }
.kpi strong { display:block; font-size:1.7rem; margin-top:.25rem; }
.bad { color:var(--rojo); font-weight:700; }
.ok { color:#067647; font-weight:700; }
.card { background:#fff; border-radius:12px; padding:1.25rem 1.35rem; margin:1rem 0; box-shadow:0 2px 10px rgba(0,0,0,.06); }
table { border-collapse:collapse; width:100%; font-size:.92rem; }
th, td { border-bottom:1px solid #d5deea; padding:.42rem .45rem; text-align:left; vertical-align:top; }
th { background:#eef3f8; color:var(--azul); font-weight:600; }
h1 { margin:0; font-size:1.55rem; }
h2, h3 { color:var(--azul); margin:0 0 .6rem; }
.note { color:#4b5563; line-height:1.5; }
.toc { display:flex; flex-wrap:wrap; gap:.45rem; margin:.8rem 0 0; }
.toc a { background:#fff; border:1px solid #c9d6e5; color:var(--azul); text-decoration:none; padding:.28rem .65rem; border-radius:999px; font-size:.82rem; }
a.btn, button.btn { display:inline-block; background:var(--azul); color:#fff; padding:.5rem .95rem; border-radius:8px; text-decoration:none; margin:.25rem .4rem .25rem 0; font-size:.92rem; border:0; }
a.btn.sec, button.btn.sec { background:#fff; color:var(--azul); border:1px solid var(--azul); }
.bar { height:8px; background:#e5edf5; border-radius:99px; min-width:70px; }
.bar > i { display:block; height:8px; background:#c4320a; border-radius:99px; }
.fac { background:#fff; padding:1rem 1.15rem; border-radius:12px; margin:1rem 0; box-shadow:0 2px 10px rgba(0,0,0,.05); page-break-inside:avoid; }
.search { flex:1; min-width:180px; max-width:520px; padding:.5rem .7rem; border:1px solid #c9d6e5; border-radius:8px; font-size:1rem; }
.muted { color:#6b7280; font-weight:400; font-size:.92rem; }
nav.sticky { display:flex; flex-wrap:wrap; gap:.45rem; padding:.7rem 2rem; background:#e8eef5; }
nav.sticky a { background:#fff; padding:.28rem .65rem; border-radius:999px; text-decoration:none; color:var(--azul); font-size:.82rem; }
.barra { position:sticky; top:0; z-index:30; background:#dfe8f2; padding:.55rem 2rem; display:flex; flex-wrap:wrap; gap:.5rem; align-items:center; box-shadow:0 2px 10px rgba(0,0,0,.08); }
.barra button, .barra a.btn { margin:0; cursor:pointer; border:0; font:inherit; }
.barra label { display:flex; flex-direction:column; font-size:.72rem; color:var(--azul); font-weight:700; letter-spacing:.02em; }
select.sel { min-width:220px; max-width:380px; padding:.4rem .5rem; font:inherit; font-weight:400; border:1px solid #c9d6e5; border-radius:8px; background:#fff; }
.aviso-elige { background:#fff7ed; border:1px solid #fdba74; border-radius:12px; padding:1rem 1.2rem; margin:1rem 0; }
.fab { position:fixed; right:1.15rem; bottom:1.15rem; z-index:40; background:var(--azul); color:#fff; border:0; border-radius:999px; padding:.7rem 1.05rem; font:inherit; font-weight:600; cursor:pointer; box-shadow:0 6px 18px rgba(0,59,112,.35); }
.prog { margin:1rem 0 1.2rem; }
.fac, .prog, #anexo, #top { scroll-margin-top: 4.6rem; }
.conteo { color:var(--azul); font-weight:600; font-size:.9rem; }
@media print {
  nav.sticky, .noprint, .search, .barra, .fab { display:none !important; }
  body { background:#fff; }
  header { -webkit-print-color-adjust:exact; print-color-adjust:exact; }
  .card, .fac { box-shadow:none; border:1px solid #d5deea; }
  a { color:inherit; text-decoration:none; }
}
"""


def html_bloques_correos(sin: list[dict[str, Any]], orden_fac: list[str] | None = None) -> tuple[str, str, str]:
    by: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    for r in sin:
        by[escuela_de(r)][programa_de(r)].append(r)
    if orden_fac is None:
        orden_fac = sorted(by.keys(), key=lambda e: (-sum(len(v) for v in by[e].values()), e))
    nav = "".join(
        f'<a href="#f-{slug_id(esc)}" onclick="elegirFacultad(\'f-{slug_id(esc)}\')">'
        f"{html.escape(esc)} ({sum(len(v) for v in by[esc].values())})</a>"
        for esc in orden_fac
        if esc in by
    )
    opciones = "".join(
        f'<option value="f-{slug_id(esc)}">{html.escape(esc)} '
        f"({sum(len(v) for v in by[esc].values())} pendientes)</option>"
        for esc in orden_fac
        if esc in by
    )
    partes = []
    for esc in orden_fac:
        progs = by.get(esc)
        if not progs:
            continue
        n_esc = sum(len(v) for v in progs.values())
        fid = f"f-{slug_id(esc)}"
        bloques_p = []
        for prog, filas in sorted(progs.items(), key=lambda kv: (-len(kv[1]), kv[0])):
            filas = sorted(filas, key=lambda r: (nombre_completo(r).lower(), r.get("correo") or ""))
            tr = "".join(
                "<tr>"
                f"<td>{html.escape(str(r.get('correo') or ''))}</td>"
                f"<td>{html.escape(nombre_completo(r))}</td>"
                f"<td>{html.escape(str(r.get('a_estado') or ''))}</td>"
                f"<td>{html.escape(etiqueta_ingreso(str(r.get('g_ultimo_ingreso') or '')))}</td>"
                "</tr>"
                for r in filas
            )
            bloques_p.append(
                f"<div class='prog' id='p-{slug_id(esc + '|' + prog)}' "
                f"data-facultad=\"{html.escape(esc, quote=True)}\" "
                f"data-programa=\"{html.escape(prog, quote=True)}\">"
                f"<h3>{html.escape(prog)} <span class='muted'>— {len(filas)} correos</span></h3>"
                "<table><thead><tr><th>Correo institucional</th><th>Estudiante</th>"
                "<th>Estado académico</th><th>Último ingreso a Google</th></tr></thead>"
                f"<tbody>{tr}</tbody></table></div>"
            )
        partes.append(
            f'<section class="fac" id="{fid}" hidden '
            f'data-facultad="{html.escape(esc, quote=True)}">'
            f"<h2>{html.escape(esc)}</h2>"
            f"<p class='note'><strong>{n_esc}</strong> estudiantes vigentes sin 2FA en esta facultad. "
            f"<button type='button' class='btn' onclick=\"descargarSeccion(document.getElementById('{fid}'))\">"
            f"Descargar CSV de esta facultad</button></p>"
            f"{''.join(bloques_p)}</section>"
        )
    return nav, "".join(partes), opciones


def js_navegacion() -> str:
    return r"""
function irArriba(){ window.scrollTo({top:0, behavior:'smooth'}); }
function selFac(){ return document.getElementById('sel-fac'); }
function selProg(){ return document.getElementById('sel-prog'); }
function textoQ(){ return ((document.getElementById('q')||{}).value || '').toLowerCase().trim(); }
function llenarProgramas(facId){
  var sel = selProg();
  if (!sel) return;
  sel.innerHTML = '<option value="">Todos los programas de esta facultad</option>';
  sel.disabled = !facId;
  if (!facId) return;
  var sec = document.getElementById(facId);
  if (!sec) return;
  sec.querySelectorAll('.prog').forEach(function(box){
    var o = document.createElement('option');
    o.value = box.id;
    o.textContent = (box.getAttribute('data-programa') || box.id) + ' (' + box.querySelectorAll('tbody tr').length + ')';
    sel.appendChild(o);
  });
}
function elegirFacultad(facId){
  var s = selFac();
  if (s) s.value = facId || '';
  llenarProgramas(facId);
  if (selProg()) selProg().value = '';
  aplicarVista(true);
}
function elegirPrograma(){
  aplicarVista(true);
}
function irAPrograma(facId, progId){
  var s = selFac();
  if (s) s.value = facId || '';
  llenarProgramas(facId);
  if (selProg()) selProg().value = progId || '';
  aplicarVista(true);
}
function aplicarVista(scroll){
  var facId = selFac() ? selFac().value : '';
  var progId = selProg() ? selProg().value : '';
  var q = textoQ();
  var total = 0;
  var primero = null;
  document.querySelectorAll('section.fac').forEach(function(sec){
    if (!facId || sec.id !== facId){
      sec.hidden = true;
      return;
    }
    var vis = 0;
    sec.querySelectorAll('.prog').forEach(function(box){
      if (progId && box.id !== progId){
        box.style.display = 'none';
        return;
      }
      var h = box.querySelector('h3');
      var n = 0;
      box.querySelectorAll('tbody tr').forEach(function(tr){
        var ok = !q || tr.innerText.toLowerCase().includes(q) || (h && h.innerText.toLowerCase().includes(q));
        tr.style.display = ok ? '' : 'none';
        if (ok) n++;
      });
      box.style.display = n ? '' : 'none';
      vis += n;
    });
    sec.hidden = vis === 0 && !!q;
    if (!q) sec.hidden = false;
    total += vis;
    if (vis && !primero) primero = sec;
  });
  document.querySelectorAll('#tabla-prog tr[data-fac-id]').forEach(function(tr){
    tr.style.display = (!facId || tr.getAttribute('data-fac-id') === facId) ? '' : 'none';
  });
  var aviso = document.getElementById('aviso-elige');
  if (aviso) aviso.hidden = !!facId;
  var nEl = document.getElementById('nfiltro');
  if (nEl){
    if (!facId) nEl.textContent = 'Elija una facultad';
    else nEl.textContent = total + ' correo' + (total===1?'':'s') + ' de esta facultad';
  }
  var btnL = document.getElementById('btn-limpiar');
  if (btnL) btnL.hidden = !q && !progId;
  var btnD = document.getElementById('btn-descarga');
  if (btnD) btnD.disabled = !facId;
  if (scroll && facId){
    var dest = document.getElementById(progId || facId);
    if (dest && !dest.hidden) dest.scrollIntoView({behavior:'smooth', block:'start'});
  }
}
function limpiarFiltro(){
  var inp = document.getElementById('q');
  if (inp) inp.value = '';
  if (selProg()) selProg().value = '';
  aplicarVista(false);
  irArriba();
}
function filtrar(){ aplicarVista(false); }
function csvCampo(s){
  s = (s==null?'':String(s)).replace(/"/g,'""');
  if (/[",\n\r]/.test(s)) return '"'+s+'"';
  return s;
}
function filasDeSeccion(sec){
  var rows = [];
  if (!sec || sec.hidden) return rows;
  var progId = selProg() ? selProg().value : '';
  sec.querySelectorAll('.prog').forEach(function(box){
    if (box.style.display === 'none') return;
    if (progId && box.id !== progId) return;
    var fac = box.getAttribute('data-facultad') || '';
    var prog = box.getAttribute('data-programa') || '';
    box.querySelectorAll('tbody tr').forEach(function(tr){
      if (tr.style.display === 'none') return;
      var tds = tr.querySelectorAll('td');
      if (tds.length < 4) return;
      rows.push([fac, prog, tds[0].innerText.trim(), tds[1].innerText.trim(), tds[2].innerText.trim(), tds[3].innerText.trim()]);
    });
  });
  return rows;
}
function bajarCsv(rows, nombre){
  if (!rows.length){ alert('No hay correos para descargar con esta selección.'); return; }
  var head = 'facultad,programa,correo,nombres,estado_academico,ultimo_ingreso_google';
  var body = rows.map(function(r){ return r.map(csvCampo).join(','); }).join('\r\n');
  var csv = '\uFEFF'+head+'\r\n'+body;
  if (navigator.msSaveBlob) {
    navigator.msSaveBlob(new Blob([csv], {type:'text/csv;charset=utf-8'}), nombre);
    return;
  }
  var sandboxed = false;
  try { sandboxed = window.self !== window.top; } catch(e) { sandboxed = true; }
  if (!sandboxed) {
    try {
      var blob = new Blob([csv], {type:'text/csv;charset=utf-8'});
      var burl = URL.createObjectURL(blob);
      var a = document.createElement('a');
      a.href = burl;
      a.download = nombre;
      a.style.display = 'none';
      document.body.appendChild(a);
      a.click();
      a.remove();
      setTimeout(function(){ URL.revokeObjectURL(burl); }, 1500);
      return;
    } catch(e) { /* fall through to popup */ }
  }
  {
    var w = window.open('', '_blank');
    if (w) {
      w.document.open('text/html','replace');
      w.document.write('<!DOCTYPE html><html><head><meta charset="utf-8"/>'
        + '<title>' + nombre + '</title></head><body>'
        + '<h3>Archivo: ' + nombre + '</h3>'
        + '<p>Seleccione todo el texto del recuadro y pegue en Excel, '
        + 'o presione el bot\u00f3n Copiar.</p>'
        + '<button onclick="var t=document.getElementById(\'ct\');'
        + 'var r=document.createRange();r.selectNodeContents(t);'
        + 'var s=window.getSelection();s.removeAllRanges();s.addRange(r);'
        + 'document.execCommand(\'copy\');alert(\'Copiado. Pegue en Excel.\');"'
        + '>Copiar al portapapeles</button>'
        + '<pre id="ct" style="white-space:pre-wrap;font-size:12px;'
        + 'border:1px solid #ccc;padding:12px;margin-top:8px;">'
        + csv.replace(/&/g,'&amp;').replace(/</g,'&lt;')
        + '</pre></body></html>');
      w.document.close();
    } else {
      alert('El navegador bloque\u00f3 la ventana emergente. '
        + 'Permita pop-ups para este sitio e intente de nuevo.');
    }
  }
}
function slugArchivo(s){
  return (s || 'facultad').toLowerCase().replace(/[^a-z0-9áéíóúñ]+/gi,'-').replace(/^-|-$/g,'').slice(0,50) || 'facultad';
}
function descargarFiltrado(){
  var s = selFac();
  if (!s || !s.value){
    alert('Elija UNA facultad en la lista de arriba. No se descarga el archivo completo.');
    return;
  }
  var sec = document.getElementById(s.value);
  var rows = filasDeSeccion(sec);
  var fac = sec ? (sec.getAttribute('data-facultad') || s.value) : s.value;
  var p = selProg() && selProg().value ? (document.getElementById(selProg().value)||{}).getAttribute('data-programa') : '';
  var nom = 'sin_2fa_' + slugArchivo(fac) + (p ? '_' + slugArchivo(p) : '') + '.csv';
  bajarCsv(rows, nom);
}
function descargarSeccion(sec){
  if (!sec) return;
  elegirFacultad(sec.id);
  bajarCsv(filasDeSeccion(sec), 'sin_2fa_' + slugArchivo(sec.getAttribute('data-facultad')) + '.csv');
}
document.addEventListener('DOMContentLoaded', function(){
  aplicarVista(false);
});
"""


def barra_busqueda(opciones_fac: str, placeholder: str = "Filtrar correos de esa facultad…") -> str:
    ph = html.escape(placeholder)
    return f"""
<div class="barra noprint" id="barra">
  <label>Facultad
    <select id="sel-fac" class="sel" onchange="elegirFacultad(this.value)">
      <option value="">— Elija una facultad —</option>
      {opciones_fac}
    </select>
  </label>
  <label>Programa
    <select id="sel-prog" class="sel" onchange="elegirPrograma()" disabled>
      <option value="">Todos los programas de esta facultad</option>
    </select>
  </label>
  <label>Buscar dentro
    <input id="q" class="search" type="search" placeholder="{ph}" oninput="filtrar()"/>
  </label>
  <button type="button" class="btn" id="btn-descarga" disabled onclick="descargarFiltrado()">Descargar esta facultad</button>
  <button type="button" class="btn sec" id="btn-limpiar" hidden onclick="limpiarFiltro()">Limpiar búsqueda</button>
  <button type="button" class="btn sec" onclick="irArriba()">↑ Inicio</button>
  <span class="conteo" id="nfiltro">Elija una facultad</span>
</div>
<button type="button" class="fab noprint" onclick="irArriba()" title="Volver al inicio">↑ Arriba</button>
"""


def html_informe_jefa(
    path: Path,
    resumen: dict[str, Any],
    planos: list[dict[str, Any]],
    cfg: dict[str, Any] | None = None,
) -> None:
    """Un solo documento: cifras, ranking y correos pendientes agrupados."""
    t = titulos_informe(cfg)
    te = {k: html.escape(v) for k, v in t.items()}
    vigentes, fac, prog = stats_campana(planos)
    sin = [r for r in vigentes if r.get("tiene_2fa") == "NO"]
    ranking = sorted(fac.items(), key=lambda kv: (-kv[1]["sin"], kv[0]))
    max_sin = max((c["sin"] for _, c in ranking), default=0) or 1
    filas_fac = []
    for e, c in ranking:
        tot = c["cuentas"]
        cob = round(100 * c["con"] / tot, 1) if tot else 0
        ancho = round(100 * c["sin"] / max_sin, 1) if c["sin"] else 0
        href = f"f-{slug_id(e)}"
        nombre = html.escape(e)
        celda = (
            f'<a href="#{href}" onclick="elegirFacultad(\'{href}\')">{nombre}</a>'
            if c["sin"]
            else nombre
        )
        filas_fac.append(
            f"<tr data-fac-id='{href}'>"
            f"<td>{celda}</td>"
            f"<td>{tot}</td><td class='ok'>{c['con']}</td>"
            f"<td class='bad'>{c['sin']}</td><td>{cob}%</td>"
            f"<td><div class='bar'><i style='width:{ancho}%'></i></div></td>"
            "</tr>"
        )
    filas_prog = []
    for (e, p), pc in sorted(prog.items(), key=lambda kv: (-kv[1]["sin"], kv[0][0], kv[0][1])):
        if not pc["sin"]:
            continue
        fid = f"f-{slug_id(e)}"
        pid = f"p-{slug_id(e + '|' + p)}"
        filas_prog.append(
            f"<tr data-fac-id='{fid}'><td>{html.escape(e)}</td>"
            f"<td><a href='#{pid}' onclick=\"irAPrograma('{fid}','{pid}')\">{html.escape(p)}</a></td>"
            f"<td>{pc['cuentas']}</td><td class='bad'>{pc['sin']}</td></tr>"
        )
    n_sin = resumen.get("n_estudiantes_sin_2fa", len(sin))
    n_match = resumen.get("n_match_estudiante", len(vigentes))
    n_con = n_match - n_sin
    cob = resumen.get("cobertura_2fa_estudiantes", 0)
    n_sin_cta = sum(1 for r in planos if r.get("perfil") == "ACADEMICO_SIN_CUENTA_GOOGLE")
    orden_fac = [e for e, c in ranking if c["sin"]]
    nav, anexo, opciones_fac = html_bloques_correos(sin, orden_fac)
    html_doc = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{te['h1']}</title>
<style>{css_presentacion()}</style>
</head><body>
<header id="top">
<p class="marca">{te['marca']}</p>
<h1>{te['h1']}</h1>
<p>Corte {html.escape(str(resumen.get('generado', '')))}. {te['intro']}</p>
</header>
{barra_busqueda(opciones_fac)}
<main>
<section class="card">
<h2>{te['resumen']}</h2>
<p class="note">{te['texto_resumen']}</p>
<p class="note noprint"><strong>Para ver o descargar correos:</strong> elija una facultad en la barra de arriba
(o un programa). No se muestra el listado completo a propósito.</p>
<p class="noprint">
<a class="btn" href="#anexo">Ir a las facultades</a>
<a class="btn sec" href="listado_sin_2fa.html">Listado operativo</a>
<a class="btn sec" href="02_estudiantes_sin_2fa.csv">CSV de TODOS (archivo grande)</a>
</p>
</section>
<div class="kpi">
  <div>{te['kpi_universo']}<strong>{n_match}</strong></div>
  <div>{te['kpi_con']}<strong class="ok">{n_con}</strong></div>
  <div>{te['kpi_sin']}<strong class="bad">{n_sin}</strong></div>
  <div>{te['kpi_cob']}<strong>{cob}%</strong></div>
</div>
<section class="card">
<h2>{te['facultades']}</h2>
<p class="note">Clic en la facultad para bajar a sus correos. Orden: más pendientes primero.</p>
<table>
<thead><tr><th>Facultad</th><th>Vigentes</th><th>Con 2FA</th><th>Pendientes</th><th>Cobertura</th><th></th></tr></thead>
<tbody>{''.join(filas_fac) or '<tr><td colspan="6">Sin estudiantes vigentes en el cruce.</td></tr>'}</tbody>
</table>
</section>
<section class="card">
<h2>{te['programas']}</h2>
<p class="note">Clic en el programa para ir a esa lista. Solo programas con pendientes.</p>
<table id="tabla-prog">
<thead><tr><th>Facultad</th><th>Programa</th><th>Vigentes</th><th>Pendientes</th></tr></thead>
<tbody>{''.join(filas_prog) or '<tr><td colspan="4">No hay pendientes.</td></tr>'}</tbody>
</table>
</section>
<details class="card noprint">
<summary><strong>Notas</strong></summary>
<ul class="note">
<li><strong>Una facultad:</strong> elija la facultad en la lista de arriba y pulse <em>Descargar esta facultad</em>. Opcional: elija un programa para bajar solo ese.</li>
<li><strong>CSV de todos:</strong> <code>02_estudiantes_sin_2fa.csv</code> es el archivo grande; no lo use si solo necesita una facultad.</li>
<li>{n_sin_cta} registros académicos vigentes sin cuenta Google no aparecen como correo pendiente.</li>
<li>{resumen.get('n_google_sin_match', 0)} cuentas Google sin ficha de estudiante vigente no se listan aquí.</li>
</ul>
</details>
<h2 id="anexo">{te['correos']}</h2>
<p class="aviso-elige noprint" id="aviso-elige">Elija <strong>una facultad</strong> en la barra de arriba.
Hasta entonces no se muestran los correos (el listado completo es demasiado grande).</p>
<nav class="toc noprint">{nav}</nav>
{anexo or '<p>No hay estudiantes vigentes sin 2FA.</p>'}
</main>
<script>{js_navegacion()}</script>
</body></html>"""
    path.write_text(html_doc, encoding="utf-8")


def html_listado_sin_2fa(
    path: Path,
    planos: list[dict[str, Any]],
    generado: str,
    cfg: dict[str, Any] | None = None,
) -> None:
    t = titulos_informe(cfg)
    te = {k: html.escape(v) for k, v in t.items()}
    sin = [
        r
        for r in planos
        if r.get("perfil") == "ESTUDIANTE_VIGENTE" and r.get("tiene_2fa") == "NO"
    ]
    _, fac, _ = stats_campana(planos)
    orden = sorted(fac.keys(), key=lambda e: (-fac[e]["sin"], e))
    nav, partes, opciones_fac = html_bloques_correos(sin, orden)
    doc = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{te['listado_h1']}</title>
<style>{css_presentacion()}</style>
</head><body>
<header id="top">
<p class="marca">{te['marca']}</p>
<h1>{te['listado_h1']}</h1>
<p>{len(sin)} cuentas pendientes · {html.escape(generado)} ·
<a href="resumen.html">{te['volver_informe']}</a></p>
</header>
{barra_busqueda(opciones_fac)}
<nav class="sticky noprint">{nav}</nav>
<main>
<p class="aviso-elige noprint" id="aviso-elige">Elija <strong>una facultad</strong> arriba. Así ve y descarga solo esa, no las ~todas.</p>
{partes or '<p>No hay filas.</p>'}
</main>
<script>{js_navegacion()}</script>
</body></html>"""
    path.write_text(doc, encoding="utf-8")


def html_report(
    path: Path,
    resumen: dict[str, Any],
    planos: list[dict[str, Any]],
    **_ignored: Any,
) -> None:
    html_informe_jefa(path, resumen, planos)


def cruzar(
    google_path: Path,
    academico_path: Path | None,
    personal_path: Path | None,
    salida: Path,
    cfg: dict[str, Any],
    academico_dir: Path | None = None,
    curriculo_path: Path | None = None,
) -> None:
    excluir = {norm_estado(x) for x in cfg.get("estados_excluir", EXCLUIR_DEFAULT)}
    ou_personal = [str(x).lower() for x in cfg.get("ou_personal", [])]
    dias_inactiva = int(cfg.get("dias_inactiva", 90))
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    salida = salida_escribible(salida)

    g_headers, g_rows = read_csv(google_path)
    if academico_dir:
        a_headers, a_rows = cargar_academico_carpeta(academico_dir)
    elif academico_path:
        a_headers, a_rows = read_csv(academico_path)
        for row in a_rows:
            row.setdefault("_fuente_archivo", academico_path.name)
            row.setdefault("_periodo", periodo_desde_nombre(academico_path.name))
        if "_fuente_archivo" not in a_headers:
            a_headers = list(a_headers) + ["_fuente_archivo", "_periodo"]
    else:
        raise SystemExit("Falta --academico o --academico-dir")
    curric_files: list[Path] = []
    if curriculo_path and curriculo_path.exists():
        curric_files = [curriculo_path]
    elif academico_dir:
        curric_files = listar_csv_curriculo(academico_dir)
    catalogo, c_map, n_filas_curriculo = cargar_catalogo_curriculo(curric_files) if curric_files else ({}, {}, 0)
    g_map = map_columns(g_headers, GOOGLE_ALIASES)
    a_map = map_columns(a_headers, ACADEMICO_ALIASES)
    if not g_map["email"]:
        raise SystemExit(f"No hallé columna de correo en Google. Encabezados: {g_headers}")
    if not a_map["email"]:
        raise SystemExit(f"No hallé columna de correo en académico. Encabezados: {a_headers}")

    fichas: dict[str, dict[str, Any]] = {}

    def get(row: dict[str, str], col: str | None) -> str:
        return row.get(col, "") if col else ""

    for row in g_rows:
        email = norm_email(get(row, g_map["email"]))
        if not email:
            continue
        f = fichas.setdefault(email, proyecto_a_ficha(email))
        f["en_google"] = True
        f["google"] = {
            "nombre": get(row, g_map["nombre"]),
            "apellido": get(row, g_map["apellido"]),
            "estado_google": get(row, g_map["estado_google"]),
            "ou": get(row, g_map["ou"]),
            "ultimo_ingreso": get(row, g_map["ultimo_ingreso"]),
            "2fa_inscrito": get(row, g_map["2fa_inscrito"]),
            "2fa_forzado": get(row, g_map["2fa_forzado"]),
        }
        f["tiene_2fa"] = is_2fa_on(f["google"]["2fa_inscrito"], f["google"]["2fa_forzado"])
        # Conserva columnas extra del Admin por si sirven después
        f["google"]["_extra"] = {k: v for k, v in row.items() if k not in set(g_map.values()) and v}

    skipped_grad = 0
    detalle_acad: list[dict[str, Any]] = []
    for row in a_rows:
        email = norm_email(get(row, a_map["email"]))
        if not email:
            continue
        estado = norm_estado(get(row, a_map["estado"]))
        if estado in excluir:
            skipped_grad += 1
            continue
        f = fichas.setdefault(email, proyecto_a_ficha(email))
        f["en_academico"] = True
        pieza = {campo: get(row, a_map[campo]) for campo in a_map}
        pieza.pop("email", None)
        if not f.get("academico"):
            f["academico"] = {**pieza, "_todas": dict(row)}
        else:
            acc = f["academico"]
            for campo, val in pieza.items():
                acc[campo] = merge_texto(acc.get(campo, ""), val)
            todas = acc.setdefault("_todas", {})
            for k, v in row.items():
                todas[k] = merge_texto(todas.get(k, ""), v)
        det = {"correo": email, "_periodo": row.get("_periodo", ""), "_fuente_archivo": row.get("_fuente_archivo", "")}
        det.update(row)
        plan = buscar_plan(
            catalogo,
            get(row, a_map.get("cod_prog")),
            get(row, a_map.get("cod_majr")),
            get(row, a_map.get("programa")),
            get(row, a_map.get("facultad")),
        )
        if plan:
            adjuntar_curriculo(f, plan)
            det["c_periodo_vigente"] = plan.get("periodo", "")
            det["c_plan"] = plan.get("plan", "")
            det["c_programa"] = plan.get("programa", "")
        detalle_acad.append(det)

    p_map = {}
    if personal_path and personal_path.exists():
        p_headers, p_rows = read_csv(personal_path)
        p_map = map_columns(p_headers, PERSONAL_ALIASES)
        if not p_map["email"]:
            print("AVISO: CSV personal sin columna de correo; se ignora.")
        else:
            for row in p_rows:
                email = norm_email(get(row, p_map["email"]))
                if not email:
                    continue
                f = fichas.setdefault(email, proyecto_a_ficha(email))
                f["en_personal"] = True
                f["personal"] = {campo: get(row, p_map[campo]) for campo in p_map if campo != "email"}
                f["personal"]["_todas"] = dict(row)

    for f in fichas.values():
        clasificar(f, ou_personal, excluir)
        f["prioridad_2fa"] = prioridad_2fa(f, dias_inactiva, now)

    salida.mkdir(parents=True, exist_ok=True)
    planos = [flatten(f) for f in sorted(fichas.values(), key=lambda x: x["correo"])]
    fields: list[str] = []
    seen: set[str] = set()
    for r in planos:
        for k in r:
            if k not in seen:
                seen.add(k)
                fields.append(k)
    if not fields:
        fields = ["correo"]

    write_csv(salida / "00_universo.csv", planos, fields)
    write_csv(
        salida / "01_sin_2fa.csv",
        [r for r in planos if r["tiene_2fa"] == "NO"],
        fields,
    )
    campana = filas_campana_csv(planos)
    write_csv(
        salida / "02_estudiantes_sin_2fa.csv",
        campana,
        [
            "facultad",
            "programa",
            "correo",
            "nombres",
            "apellidos",
            "estado_academico",
            "ultimo_ingreso_google",
            "periodo_plan_vigente",
            "cod_prog",
        ],
    )
    write_csv(
        salida / "03_google_sin_match_academico.csv",
        [r for r in planos if r["en_google"] == "SI" and r["en_academico"] == "NO"],
        fields,
    )
    write_csv(
        salida / "04_academico_sin_cuenta_google.csv",
        [r for r in planos if r["perfil"] == "ACADEMICO_SIN_CUENTA_GOOGLE"],
        fields,
    )
    write_csv(
        salida / "05_prioridad_alta_2fa.csv",
        [r for r in planos if r["prioridad_2fa"].startswith("ALTA")],
        fields,
    )

    fac = []
    gstats = defaultdict(Counter)
    for r in planos:
        if r["en_academico"] != "SI" or r["en_google"] != "SI":
            continue
        k = escuela_de(r)
        gstats[k]["cuentas"] += 1
        if r["tiene_2fa"] == "NO":
            gstats[k]["sin_2fa"] += 1
        if r["tiene_2fa"] == "SI":
            gstats[k]["con_2fa"] += 1
    for k, c in sorted(gstats.items()):
        tot = c["cuentas"]
        fac.append(
            {
                "facultad": k,
                "cuentas": tot,
                "sin_2fa": c["sin_2fa"],
                "cobertura_2fa": round(100 * c["con_2fa"] / tot, 1) if tot else 0,
            }
        )
    write_csv(
        salida / "06_cobertura_2fa_facultad.csv",
        fac,
        ["facultad", "cuentas", "sin_2fa", "cobertura_2fa"],
    )

    prog = defaultdict(Counter)
    for r in planos:
        if r["en_academico"] != "SI" or r["en_google"] != "SI":
            continue
        k = f"{r.get('a_facultad') or '-'} | {r.get('a_programa') or '-'} | seccion {r.get('a_seccion') or '-'}"
        prog[k]["cuentas"] += 1
        if r["tiene_2fa"] == "NO":
            prog[k]["sin_2fa"] += 1
        if r["tiene_2fa"] == "SI":
            prog[k]["con_2fa"] += 1
    prog_rows = []
    for k, c in sorted(prog.items()):
        tot = c["cuentas"]
        prog_rows.append(
            {
                "facultad_programa_seccion": k,
                "cuentas": tot,
                "sin_2fa": c["sin_2fa"],
                "cobertura_2fa": round(100 * c["con_2fa"] / tot, 1) if tot else 0,
            }
        )
    write_csv(
        salida / "07_cobertura_2fa_programa_seccion.csv",
        prog_rows,
        ["facultad_programa_seccion", "cuentas", "sin_2fa", "cobertura_2fa"],
    )
    det_fields: list[str] = []
    det_seen: set[str] = set()
    for r in detalle_acad:
        for k in r:
            if k not in det_seen:
                det_seen.add(k)
                det_fields.append(k)
    if det_fields:
        write_csv(salida / "08_academico_filas.csv", detalle_acad, det_fields)

    planes_vigentes: list[dict[str, Any]] = []
    seen_plan: set[tuple] = set()
    for pieza in catalogo.values():
        ident = (
            pieza.get("periodo", ""),
            pieza.get("cod_prog", ""),
            pieza.get("major", ""),
            pieza.get("programa", ""),
        )
        if ident in seen_plan:
            continue
        seen_plan.add(ident)
        row_out = {
            "periodo_vigente": pieza.get("periodo", ""),
            "cod_esc": pieza.get("cod_esc", ""),
            "escuela": pieza.get("escuela", ""),
            "cod_prog": pieza.get("cod_prog", ""),
            "programa": pieza.get("programa", ""),
            "major": pieza.get("major", ""),
            "plan": pieza.get("plan", ""),
            "nivel": pieza.get("nivel", ""),
            "tipo": pieza.get("tipo", ""),
            "campus": pieza.get("campus", ""),
            "distrito": pieza.get("distrito", ""),
            "web": pieza.get("web", ""),
            "periodo_efectivo": pieza.get("periodo_efectivo", ""),
            "_fuente": pieza.get("_fuente", ""),
        }
        extra = pieza.get("_todas") or {}
        for k, v in extra.items():
            if k and f"{k}" not in row_out:
                row_out[k] = v
        planes_vigentes.append(row_out)
    planes_vigentes.sort(key=lambda r: (str(r.get("escuela", "")), str(r.get("programa", ""))))
    if planes_vigentes:
        plan_fields: list[str] = []
        seen_pf: set[str] = set()
        for r in planes_vigentes:
            for k in r:
                if k not in seen_pf:
                    seen_pf.add(k)
                    plan_fields.append(k)
        write_csv(salida / "09_catalogo_planes_vigentes.csv", planes_vigentes, plan_fields)

    plan_stats = defaultdict(Counter)
    for r in planos:
        if r["en_academico"] != "SI" or r["en_google"] != "SI":
            continue
        k = (
            r.get("c_escuela") or r.get("a_facultad") or "-",
            r.get("c_programa") or r.get("a_programa") or "-",
            r.get("c_periodo_vigente") or "-",
        )
        plan_stats[k]["cuentas"] += 1
        if r["tiene_2fa"] == "NO":
            plan_stats[k]["sin_2fa"] += 1
        if r["tiene_2fa"] == "SI":
            plan_stats[k]["con_2fa"] += 1
    plan_rows = []
    for (esc, prog, per), c in sorted(plan_stats.items()):
        tot = c["cuentas"]
        plan_rows.append(
            {
                "escuela": esc,
                "programa": prog,
                "periodo_vigente": per,
                "cuentas": tot,
                "sin_2fa": c["sin_2fa"],
                "cobertura_2fa": round(100 * c["con_2fa"] / tot, 1) if tot else 0,
            }
        )
    write_csv(
        salida / "10_cobertura_2fa_plan_vigente.csv",
        plan_rows,
        ["escuela", "programa", "periodo_vigente", "cuentas", "sin_2fa", "cobertura_2fa"],
    )

    n_google = sum(1 for r in planos if r["en_google"] == "SI")
    n_acad = sum(1 for r in planos if r["en_academico"] == "SI")
    n_match = sum(1 for r in planos if r["perfil"] == "ESTUDIANTE_VIGENTE")
    n_sin = sum(1 for r in planos if r["tiene_2fa"] == "NO")
    n_con = sum(1 for r in planos if r["tiene_2fa"] == "SI")
    n_est_sin = sum(1 for r in planos if r["perfil"] == "ESTUDIANTE_VIGENTE" and r["tiene_2fa"] == "NO")
    n_est_con = sum(1 for r in planos if r["perfil"] == "ESTUDIANTE_VIGENTE" and r["tiene_2fa"] == "SI")
    n_google_sin_match = sum(1 for r in planos if r["en_google"] == "SI" and r["en_academico"] == "NO")
    n_match_curr = sum(1 for r in planos if r.get("c_match") == "SI")
    cob_est = round(100 * n_est_con / n_match, 1) if n_match else 0
    cobertura = round(100 * n_con / n_google, 1) if n_google else 0
    perfiles = dict(Counter(r["perfil"] for r in planos))
    resumen = {
        "generado": datetime.now().isoformat(timespec="seconds"),
        "n_google": n_google,
        "n_academico": n_acad,
        "n_match_estudiante": n_match,
        "n_estudiantes_sin_2fa": n_est_sin,
        "cobertura_2fa_estudiantes": cob_est,
        "n_sin_2fa": n_sin,
        "cobertura_2fa": cobertura,
        "n_google_sin_match": n_google_sin_match,
        "n_planes_vigentes": len(planes_vigentes),
        "n_filas_curriculo": n_filas_curriculo,
        "n_match_curriculo": n_match_curr,
        "perfiles": perfiles,
        "graduados_filtrados_en_academico": skipped_grad,
        "mapeo_google": g_map,
        "mapeo_academico": a_map,
        "mapeo_personal": p_map,
        "mapeo_curriculo": c_map,
    }
    (salida / "resumen.json").write_text(json.dumps(resumen, ensure_ascii=False, indent=2), encoding="utf-8")
    html_informe_jefa(salida / "resumen.html", resumen, planos, cfg)
    html_listado_sin_2fa(salida / "listado_sin_2fa.html", planos, str(resumen.get("generado", "")), cfg)

    print("Listo ->", salida.resolve())
    print(f"Google: {n_google} | Académico vigente: {n_acad} | Match estudiante: {n_match}")
    print(f"Estudiantes sin 2FA: {n_est_sin} | Cobertura 2FA estudiantes: {cob_est}%")
    print(f"Sin 2FA dominio: {n_sin} | Cobertura 2FA dominio: {cobertura}%")
    print(f"Radar Google sin ficha vigente: {n_google_sin_match} | Egresados filtrados del académico: {skipped_grad}")
    print(f"Currículo: {n_filas_curriculo} filas históricas -> {len(planes_vigentes)} planes vigentes | match plan: {n_match_curr}")
    print("Informe jefa:", (salida / "resumen.html").resolve())
    print("Correos sin 2FA:", (salida / "listado_sin_2fa.html").resolve())
    print("Columnas Google detectadas:", {k: v for k, v in g_map.items() if v})
    print("Columnas académico detectadas:", {k: v for k, v in a_map.items() if v})
    if c_map:
        print("Columnas currículo detectadas:", {k: v for k, v in c_map.items() if v})


def main() -> None:
    root = Path(__file__).resolve().parent
    ap = argparse.ArgumentParser(description="Cruce Google Admin × académico (2FA y ficha institucional)")
    ap.add_argument("--google", type=Path)
    ap.add_argument("--academico", type=Path)
    ap.add_argument("--academico-dir", type=Path, help="Carpeta con varios CSV de inscritos (se unen)")
    ap.add_argument("--personal", type=Path, help="CSV opcional de docentes/administrativos (GH/nómina)")
    ap.add_argument("--curriculo", type=Path, help="CSV/XLSX Vista de currículo (si está en la carpeta se detecta solo)")
    ap.add_argument("--carpeta", type=Path, help="Una carpeta con Google + inscritos + VISTA DE CURRICULO (xlsx/csv)")
    ap.add_argument("--salida", type=Path)
    ap.add_argument("--config", type=Path, default=root / "config.yaml")
    ap.add_argument("--ejemplo", action="store_true", help="Corre con CSV de entrada/_ejemplos")
    ap.add_argument("--inspeccionar", type=Path, help="Solo analiza encabezados de una carpeta (sin cruce)")
    args = ap.parse_args()
    cfg = load_config(args.config if args.config.exists() else root / "config.example.yaml")

    if args.inspeccionar:
        inspeccionar_academico(args.inspeccionar)
        return
    if args.ejemplo:
        base = root / "entrada" / "_ejemplos"
        cruzar(
            base / "google_admin.csv",
            base / "academico.csv",
            base / "personal.csv",
            args.salida or root / "salida",
            cfg,
            curriculo_path=base / "curriculo.csv" if (base / "curriculo.csv").exists() else None,
        )
        return
    if args.carpeta:
        fuentes = descubrir_fuentes(args.carpeta)
        if not fuentes["google"]:
            raise SystemExit(f"No hallé export Google (User_Download*.csv) en {args.carpeta}")
        if not fuentes["academico"]:
            raise SystemExit(f"No hallé CSV de inscritos en {args.carpeta}")
        cruzar(
            fuentes["google"][0],
            None,
            args.personal,
            args.salida or root / "salida",
            cfg,
            academico_dir=args.carpeta,
            curriculo_path=fuentes["curriculo"][0] if fuentes["curriculo"] else None,
        )
        return
    if not args.google or not (args.academico or args.academico_dir):
        ap.error("Indica --carpeta, o --google y --academico/--academico-dir, o --ejemplo / --inspeccionar")
    cruzar(
        args.google,
        args.academico,
        args.personal,
        args.salida or root / "salida",
        cfg,
        academico_dir=args.academico_dir,
        curriculo_path=args.curriculo,
    )


if __name__ == "__main__":
    main()
